from __future__ import annotations

import csv
import re
from dataclasses import dataclass
from io import BytesIO, StringIO
from pathlib import Path

from docx import Document as DocxDocument
from docx.oxml.table import CT_Tbl
from docx.oxml.text.paragraph import CT_P
from docx.table import Table
from docx.text.paragraph import Paragraph
from openpyxl import load_workbook
from pypdf import PdfReader

from .models import ExtractionStatus

MAX_ATTACHMENT_SIZE_BYTES = 20 * 1024 * 1024
ALLOWED_EXTENSIONS = {".docx", ".xlsx", ".csv", ".tsv", ".txt", ".pdf"}

EXPECTED_MIME_TYPES = {
    ".docx": {
        "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        "application/octet-stream",
    },
    ".xlsx": {
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/octet-stream",
    },
    ".csv": {"text/csv", "application/csv", "application/octet-stream"},
    ".tsv": {"text/tab-separated-values", "text/plain", "application/octet-stream"},
    ".txt": {"text/plain", "application/octet-stream"},
    ".pdf": {"application/pdf", "application/octet-stream"},
}


class AttachmentValidationError(Exception):
    """Raised when uploaded attachment does not satisfy policy checks."""


@dataclass(slots=True)
class ParsedAttachment:
    extracted_text: str
    extraction_status: ExtractionStatus
    extraction_error: str



def validate_attachment(filename: str, mime_type: str, size_bytes: int) -> None:
    extension = Path(filename).suffix.lower()
    if extension not in ALLOWED_EXTENSIONS:
        supported = ", ".join(sorted(ALLOWED_EXTENSIONS))
        raise AttachmentValidationError(
            f"Unsupported extension '{extension}'. Supported: {supported}"
        )

    if len(Path(filename).name) > 255:
        raise AttachmentValidationError("Filename is too long")

    if size_bytes <= 0:
        raise AttachmentValidationError("Attachment is empty")

    if size_bytes > MAX_ATTACHMENT_SIZE_BYTES:
        raise AttachmentValidationError(
            f"Attachment too large ({size_bytes} bytes). Limit: {MAX_ATTACHMENT_SIZE_BYTES} bytes"
        )

    expected = EXPECTED_MIME_TYPES.get(extension, set())
    if mime_type and expected and mime_type not in expected:
        raise AttachmentValidationError(
            f"Invalid MIME type '{mime_type}' for extension '{extension}'"
        )


def validate_attachment_content(filename: str, raw_bytes: bytes) -> None:
    extension = Path(filename).suffix.lower()
    if extension == ".pdf" and not raw_bytes.startswith(b"%PDF"):
        raise AttachmentValidationError("Invalid PDF signature")
    if extension in {".docx", ".xlsx"} and not _looks_like_zip(raw_bytes):
        raise AttachmentValidationError(
            f"Invalid binary signature for extension '{extension}'"
        )
    if extension in {".txt", ".csv", ".tsv"} and b"\x00" in raw_bytes:
        raise AttachmentValidationError("Text-based attachment contains binary null bytes")


def parse_attachment(filename: str, raw_bytes: bytes) -> ParsedAttachment:
    extension = Path(filename).suffix.lower()
    try:
        if extension == ".txt":
            text = _parse_text(raw_bytes)
            return _build_result(text)
        if extension in {".csv", ".tsv"}:
            text = _parse_delimited(raw_bytes, delimiter="," if extension == ".csv" else "\t")
            return _build_result(text)
        if extension == ".docx":
            text = _parse_docx(raw_bytes)
            return _build_result(text)
        if extension == ".xlsx":
            text = _parse_xlsx(raw_bytes)
            return _build_result(text)
        if extension == ".pdf":
            text, is_partial = _parse_pdf(raw_bytes)
            return _build_result(text, is_partial=is_partial)
    except Exception as exc:
        return ParsedAttachment(
            extracted_text="",
            extraction_status=ExtractionStatus.FAILED,
            extraction_error=_normalize_error(exc),
        )

    return ParsedAttachment(
        extracted_text="",
        extraction_status=ExtractionStatus.FAILED,
        extraction_error=f"Unsupported extension '{extension}'",
    )


def _build_result(text: str, *, is_partial: bool = False) -> ParsedAttachment:
    normalized = text.strip()
    if not normalized:
        return ParsedAttachment(
            extracted_text="",
            extraction_status=ExtractionStatus.PARTIAL,
            extraction_error="No extractable text was found",
        )
    status = ExtractionStatus.PARTIAL if is_partial else ExtractionStatus.OK
    return ParsedAttachment(
        extracted_text=normalized,
        extraction_status=status,
        extraction_error="" if status == ExtractionStatus.OK else "Some content could not be extracted",
    )


def _parse_text(raw_bytes: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            return raw_bytes.decode(encoding)
        except UnicodeDecodeError:
            continue
    return raw_bytes.decode("utf-8", errors="replace")


def _parse_delimited(raw_bytes: bytes, *, delimiter: str) -> str:
    decoded = _parse_text(raw_bytes)
    reader = csv.reader(StringIO(decoded), delimiter=delimiter)
    lines: list[str] = []
    for row in reader:
        cells = [str(cell).strip() for cell in row]
        line = "\t".join(cell for cell in cells if cell)
        if line:
            lines.append(line)
    return "\n".join(lines)


def _parse_docx(raw_bytes: bytes) -> str:
    document = DocxDocument(BytesIO(raw_bytes))
    blocks: list[str] = []
    pending_list_lines: list[str] = []

    def _flush_list() -> None:
        nonlocal pending_list_lines
        if not pending_list_lines:
            return
        blocks.append("\n".join(pending_list_lines))
        pending_list_lines = []

    for child in document.element.body.iterchildren():
        if isinstance(child, CT_P):
            paragraph = Paragraph(child, document)
            paragraph_line = _parse_docx_paragraph(paragraph)
            if not paragraph_line:
                continue
            if paragraph_line.startswith("- ") or paragraph_line.startswith("1. ") or paragraph_line.startswith("  - "):
                pending_list_lines.append(paragraph_line)
            else:
                _flush_list()
                blocks.append(paragraph_line)
            continue
        if isinstance(child, CT_Tbl):
            _flush_list()
            table = Table(child, document)
            table_markdown = _parse_docx_table(table)
            if table_markdown:
                blocks.append(table_markdown)

    _flush_list()

    return "\n\n".join(blocks)


def _parse_docx_paragraph(paragraph: Paragraph) -> str:
    text = re.sub(r"\s+", " ", (paragraph.text or "").strip())
    if not text:
        return ""

    style_name = re.sub(r"\s+", " ", str(getattr(paragraph.style, "name", "") or "")).strip().lower()
    heading_match = re.search(r"\bheading\s*([1-6])\b", style_name)
    if heading_match:
        level = max(1, min(6, int(heading_match.group(1))))
        return f"{'#' * level} {text}"

    if _is_list_like_paragraph(paragraph, style_name):
        level = _list_level(paragraph, style_name)
        marker = "1." if "number" in style_name else "-"
        return f"{'  ' * level}{marker} {text}"

    return text


def _is_list_like_paragraph(paragraph: Paragraph, style_name: str) -> bool:
    if "list bullet" in style_name or "list number" in style_name:
        return True
    ppr = getattr(paragraph._p, "pPr", None)
    if ppr is None:
        return False
    return getattr(ppr, "numPr", None) is not None


def _list_level(paragraph: Paragraph, style_name: str) -> int:
    style_match = re.search(r"\blist\s+(?:bullet|number)\s*(\d+)\b", style_name)
    if style_match:
        return max(0, min(4, int(style_match.group(1)) - 1))
    ppr = getattr(paragraph._p, "pPr", None)
    if ppr is not None and getattr(ppr, "numPr", None) is not None:
        ilvl = getattr(ppr.numPr, "ilvl", None)
        if ilvl is not None and getattr(ilvl, "val", None) is not None:
            try:
                return max(0, min(4, int(ilvl.val)))
            except (TypeError, ValueError):
                return 0
    return 0


def _parse_docx_table(table: Table) -> str:
    rows: list[list[str]] = []
    for row in table.rows:
        cells = []
        for cell in row.cells:
            normalized = re.sub(r"\s+", " ", (cell.text or "").strip())
            normalized = normalized.replace("|", "\\|")
            cells.append(normalized)
        if any(cell for cell in cells):
            rows.append(cells)
    if not rows:
        return ""
    width = max(len(row) for row in rows)
    normalized_rows = [row + [""] * (width - len(row)) for row in rows]
    header = normalized_rows[0]
    separator = ["---"] * width
    lines = [
        f"| {' | '.join(header)} |",
        f"| {' | '.join(separator)} |",
    ]
    for row in normalized_rows[1:]:
        lines.append(f"| {' | '.join(row)} |")
    return "\n".join(lines)


def _parse_xlsx(raw_bytes: bytes) -> str:
    workbook = load_workbook(filename=BytesIO(raw_bytes), data_only=True, read_only=True)
    lines: list[str] = []
    try:
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            lines.append(f"# Sheet: {sheet_name}")
            for row in sheet.iter_rows(values_only=True):
                cells = [str(cell).strip() for cell in row if cell is not None and str(cell).strip()]
                if cells:
                    lines.append("\t".join(cells))
    finally:
        workbook.close()
    return "\n".join(lines)


def _parse_pdf(raw_bytes: bytes) -> tuple[str, bool]:
    reader = PdfReader(BytesIO(raw_bytes))
    pages: list[str] = []
    missing_pages = 0
    for page in reader.pages:
        text = (page.extract_text() or "").strip()
        if text:
            pages.append(text)
        else:
            missing_pages += 1
    return "\n\n".join(pages), missing_pages > 0


def _normalize_error(exc: Exception) -> str:
    message = str(exc).strip() or exc.__class__.__name__
    # Keep API-friendly one-line messages.
    return re.sub(r"\s+", " ", message)


def _looks_like_zip(raw_bytes: bytes) -> bool:
    return raw_bytes.startswith((b"PK\x03\x04", b"PK\x05\x06", b"PK\x07\x08"))
